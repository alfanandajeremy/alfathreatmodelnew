import streamlit as st
import groq
from PIL import Image
import time
import base64
import re

# HTML/Markdown -> PDF
import markdown2
from xhtml2pdf import pisa
from io import BytesIO, StringIO

# Excel
import pandas as pd

# =========================
# 1) Konfigurasi Halaman
# =========================
st.set_page_config(
    page_title="Alfa Threat Model Expert Analysis",
    page_icon="⚡️",
    layout="centered"
)

# =========================
# 2) PDF Utilities (rapi)
# =========================
def _shape_tables_for_pdf(raw_html: str) -> str:
    """
    Post-process HTML agar tabel lebih rapi untuk xhtml2pdf:
    - Tambahkan <thead> pada baris header
    - Tambahkan <colgroup> dengan lebar kolom yang presisi jika header cocok
    - Rapikan sel & mencegah baris terbelah
    """
    def strip_tags(t):
        return re.sub(r"<[^>]+>", "", t, flags=re.I | re.S).strip()

    def add_colgroup_for_known_table(table_html: str) -> str:
        m = re.search(r"<tr[^>]*>(.*?)</tr>", table_html, flags=re.I | re.S)
        if not m:
            return table_html
        header_row = m.group(1)
        headers = re.findall(r"<t[hd][^>]*>(.*?)</t[hd]>", header_row, flags=re.I | re.S)
        headers_text = [strip_tags(h).lower() for h in headers]
        if not headers_text:
            return table_html

        def norm(h):
            h = h.replace("authentication", "au").replace("availability", "av")
            h = h.replace("confidentiality", "c").replace("integrity", "i")
            h = h.replace("non-repudiation", "n")
            return h.strip().lower()

        mapped = [norm(h) for h in headers_text]

        score = 0
        if any(x.startswith("flow") for x in mapped): score += 1
        if "threat" in "".join(mapped): score += 1
        for k in ["c", "i", "au", "av", "n"]:
            if k in mapped: score += 1
        if any("scenario" in x for x in mapped): score += 1
        if any("rekomendasi" in x for x in mapped): score += 1
        if score < max(5, len(mapped) // 2):
            return table_html  # bukan tabel target

        widths = [18, 22, 4, 4, 4, 4, 4, 22, 18]
        n = len(mapped)
        if n == 9:
            weights = widths
        else:
            base = [20, 20] + [int(60 / max(1, n - 2))] * (n - 2)
            weights = base

        colgroup = "<colgroup>" + "".join([f'<col style="width:{w}%;"/>' for w in weights]) + "</colgroup>"
        table_html = re.sub(r"(<table[^>]*>)", r"\1" + colgroup, table_html, flags=re.I | re.S, count=1)
        return table_html

    def ensure_thead(table_html: str) -> str:
        if re.search(r"<thead", table_html, flags=re.I):
            return table_html
        return re.sub(
            r"(<table[^>]*>\s*)(<tr[^>]*>.*?</tr>)",
            r"\1<thead>\2</thead>",
            table_html,
            flags=re.I | re.S,
            count=1
        )

    def process_table(m):
        tbl = m.group(0)
        tbl = ensure_thead(tbl)
        tbl = add_colgroup_for_known_table(tbl)
        return tbl

    shaped = re.sub(r"<table[^>]*>.*?</table>", process_table, raw_html, flags=re.I | re.S)
    return shaped


def create_pdf_with_xhtml2pdf(markdown_content, filename="alfa_threat_analysis.pdf"):
    """
    PDF generator dengan styling tabel yang sudah dibereskan:
    - Header tabel diulang (thead)
    - Lebar kolom diatur (colgroup)
    - Wrap teks, border rapi, baris tidak terbelah
    """
    try:
        html_string = markdown2.markdown(
            markdown_content,
            extras=["tables", "fenced-code-blocks", "code-friendly"]
        )
        html_string = _shape_tables_for_pdf(html_string)

        full_html = f"""
        <html>
            <head>
                <meta charset="UTF-8">
                <style>
                    @page {{ margin: 1.3cm; }}
                    body {{
                        font-family: Helvetica, Arial, sans-serif;
                        font-size: 10pt;
                        line-height: 1.35;
                    }}
                    h1, h2, h3, h4 {{ margin: 8px 0 6px; font-weight: bold; }}
                    p {{ margin: 4px 0; }}

                    table {{
                        border-collapse: collapse;
                        width: 100%;
                        table-layout: fixed;
                        margin: 8px 0;
                        page-break-inside: auto;
                        -pdf-keep-in-frame: auto;
                    }}
                    thead {{ display: table-header-group; }}
                    tr {{ page-break-inside: avoid; }}
                    th, td {{
                        border: 0.8pt solid #444;
                        padding: 6px 6px;
                        vertical-align: top;
                        word-wrap: break-word;
                        white-space: pre-wrap;
                        font-size: 9pt;
                    }}
                    th {{
                        background-color: #f0f0f0;
                        font-weight: bold;
                        text-align: center;
                    }}
                    /* Pusatkan kolom CIA/Au/Av/N jika ada */
                    th:nth-child(3), th:nth-child(4), th:nth-child(5), th:nth-child(6), th:nth-child(7),
                    td:nth-child(3), td:nth-child(4), td:nth-child(5), td:nth-child(6), td:nth-child(7) {{
                        text-align: center;
                        width: 4%;
                    }}
                    pre, code {{
                        background-color: #f4f4f4;
                        padding: 2px 4px;
                        border: 1px solid #ddd;
                        border-radius: 3px;
                        font-family: 'Courier New', monospace;
                        white-space: pre-wrap;
                        word-wrap: break-word;
                        font-size: 8pt;
                    }}
                </style>
            </head>
            <body>
                {html_string}
            </body>
        </html>
        """

        result_file = BytesIO()
        pisa_status = pisa.CreatePDF(
            BytesIO(full_html.encode("UTF-8")),
            dest=result_file,
            encoding="UTF-8"
        )
        if pisa_status.err:
            st.error(f"Gagal membuat PDF: {pisa_status.err}")
            return ""

        result_file.seek(0)
        b64_pdf = base64.b64encode(result_file.read()).decode()
        href = f'''
        <a href="data:application/pdf;base64,{b64_pdf}"
           download="{filename}"
           style="display:inline-block; padding:8px 12px; background:#FF4B4B; color:#fff;
                  border-radius:5px; text-decoration:none; font-weight:bold;">
           💾 Unduh Analisis (PDF)
        </a>
        '''
        return href

    except Exception as e:
        st.error(f"Terjadi kesalahan saat membuat PDF: {e}")
        return ""


# =========================
# 3) Excel Utilities (super rapi)
# =========================
def _clean_sheet_name(name: str) -> str:
    cleaned = re.sub(r'[:\\/\?\*\[\]]', ' ', name).strip()
    return (cleaned[:31] or "Sheet") if cleaned else "Sheet"

def _openpyxl_style_table(ws, df):
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.formatting.rule import ColorScaleRule

    max_row = ws.max_row
    max_col = ws.max_column
    if max_row < 1 or max_col < 1:
        return

    ref = f"A1:{get_column_letter(max_col)}{max_row}"
    try:
        t = Table(displayName=f"T_{ws.title.replace(' ', '_')}", ref=ref)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        t.tableStyleInfo = style
        ws.add_table(t)
    except Exception:
        pass

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ref

    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(val), 80))
        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 60)

    numeric_cols = []
    for j, col in enumerate(df.columns, start=1):
        series = pd.to_numeric(df[col], errors='coerce')
        if series.notna().sum() >= max(1, len(series) // 2):
            numeric_cols.append(j)
    if numeric_cols:
        for j in numeric_cols:
            col_letter = get_column_letter(j)
            rng = f"{col_letter}2:{col_letter}{max_row}"
            try:
                rule = ColorScaleRule(start_type="min", mid_type="percentile", end_type="max",
                                      mid_value=50)
                ws.conditional_formatting.add(rng, rule)
            except Exception:
                pass

    try:
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_title_rows = "1:1"
        ws.page_margins.left = ws.page_margins.right = 0.4
        ws.page_margins.top = ws.page_margins.bottom = 0.5
    except Exception:
        pass

def _xlsxwriter_style_table(workbook, worksheet, df):
    wrap = workbook.add_format({'text_wrap': True, 'valign': 'top'})
    worksheet.set_column(0, len(df.columns) - 1, 12, wrap)
    worksheet.freeze_panes(1, 0)

    rows = len(df.index)
    cols = len(df.columns)

    try:
        worksheet.add_table(0, 0, rows, cols - 1, {
            'columns': [{'header': str(h)} for h in df.columns],
            'style': 'Table Style Medium 9',
            'autofilter': True
        })
    except Exception:
        pass

    for i, col in enumerate(df.columns):
        col_values = [str(col)] + ["" if pd.isna(x) else str(x) for x in df[col].tolist()]
        max_len = min(max(len(x) for x in col_values), 80)
        worksheet.set_column(i, i, min(max(12, max_len + 2), 60), wrap)

    for i, col in enumerate(df.columns):
        series = pd.to_numeric(df[col], errors='coerce')
        if series.notna().sum() >= max(1, len(series)//2):
            worksheet.conditional_format(1, i, rows, i, {'type': '3_color_scale'})

    worksheet.set_landscape()
    worksheet.fit_to_pages(1, 0)
    worksheet.set_margins(left=0.4, right=0.4, top=0.5, bottom=0.5)

def _parse_markdown_tables_simple(md_text):
    """
    Fallback sederhana: parse tabel Markdown (| col | ...).
    Kembalikan list DataFrame.
    """
    lines = md_text.strip().splitlines()
    tables = []
    i = 0
    while i < len(lines):
        if '|' in lines[i]:
            if i + 1 < len(lines) and re.search(r'^\s*\|?\s*:?-{3,}.*\|.*-+', lines[i+1]):
                start = i
                j = i
                while j < len(lines) and '|' in lines[j] and not lines[j].strip().startswith('```'):
                    j += 1
                block = lines[start:j]

                def split_row(row):
                    row = row.strip()
                    if row.startswith('|'): row = row[1:]
                    if row.endswith('|'): row = row[:-1]
                    return [c.strip() for c in row.split('|')]

                if len(block) >= 2:
                    header = split_row(block[0])
                    data_rows = [split_row(r) for r in block[2:]]
                    max_cols = len(header)
                    norm = []
                    for r in data_rows:
                        if len(r) < max_cols:
                            r = r + ['']*(max_cols-len(r))
                        elif len(r) > max_cols:
                            r = r[:max_cols]
                        norm.append(r)
                    try:
                        df = pd.DataFrame(norm, columns=header)
                        tables.append(df)
                    except Exception:
                        pass
                i = j
                continue
        i += 1
    return tables

def create_excel_from_markdown(markdown_content: str, filename: str = "alfa_threat_analysis.xlsx"):
    """
    Mengubah output markdown asisten menjadi file Excel yang rapi:
    - Tiap tabel → sheet sebagai Excel Table (banded rows, filter)
    - Tidak ada tabel → sheet 'Output' berisi teks (1 kolom)
    - Freeze header, wrap text, auto-fit kolom, print setup, conditional formatting numerik
    """
    try:
        html_string = markdown2.markdown(
            markdown_content,
            extras=["tables", "fenced-code-blocks", "code-friendly"]
        )
        tables = []
        try:
            tables = pd.read_html(StringIO(html_string))
        except Exception:
            tables = []

        if not tables:
            tables = _parse_markdown_tables_simple(markdown_content)

        only_text_mode = False
        if not tables:
            only_text_mode = True
            tables = [pd.DataFrame({"Output": [markdown_content]})]

        output = BytesIO()

        engine = None
        try:
            import openpyxl  # noqa
            engine = "openpyxl"
        except Exception:
            try:
                import xlsxwriter  # noqa
                engine = "xlsxwriter"
            except Exception:
                engine = None

        if not engine:
            st.error("Tidak ditemukan engine Excel (openpyxl/xlsxwriter). Tambahkan salah satunya ke environment.")
            return None

        if engine == "openpyxl":
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                for idx, df in enumerate(tables, start=1):
                    df = df.applymap(lambda x: str(x).strip() if pd.notna(x) else x)
                    sheet_name = "Output" if (only_text_mode and idx == 1) else f"Table{idx}"
                    sheet_name = _clean_sheet_name(sheet_name)
                    df.to_excel(writer, index=False, sheet_name=sheet_name)
                    ws = writer.sheets[sheet_name]
                    _openpyxl_style_table(ws, df)
        else:
            with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
                for idx, df in enumerate(tables, start=1):
                    df = df.applymap(lambda x: str(x).strip() if pd.notna(x) else x)
                    sheet_name = "Output" if (only_text_mode and idx == 1) else f"Table{idx}"
                    sheet_name = _clean_sheet_name(sheet_name)
                    df.to_excel(writer, index=False, sheet_name=sheet_name)
                    ws = writer.sheets[sheet_name]
                    _xlsxwriter_style_table(writer.book, ws, df)

        output.seek(0)
        return output.read()
    except Exception as e:
        st.error(f"Gagal membuat Excel: {e}")
        return None

# =========================
# 4) Main App
# =========================
def main_app():
    st.session_state.last_activity = time.time()

    st.title("⚡️ Alfa Threat Model Expert Analysis")
    st.write("Deskripsikan Flow aplikasi maka AlfaThreat akan menganalisa-nya dengan mudah dan akurat.")
    st.markdown("---")

    SYSTEM_PROMPT_CONTENT = """
    Anda adalah pakar keamanan siber dengan pengalaman 30 tahun, yang mahir dalam mengidentifikasi dan menganalisis potensi ancaman keamanan siber berdasarkan ASVS dan CWE. Jika ada pertanyaan di luar cyber security jangan berikan jawaban.
    Peran Anda adalah membuat kajian/report dalam tabel yang rapi serta up to date dengan perkembangan cyber security atas poin berikut:
    1. Gambarkan threat modelling diagram DFD menggunakan arrow.

    2. Berikan dalam tabel dan mapping nomor CWE-nya berdasarkan threat.

    3. Berikan list kerentanan dalam satu tabel pada Confidentiality, Integrity, Authentication, Availability, Non-repudiation atas threat tersebut termasuk dalam threat apa (contoh: SQLi, XSS, IDOR, DDoS). Lalu pada kolom sampingnya berikan Scenario serangan, lalu pada kolom sampingnya berikan rekomendasi keamanannya (minimal 5 rekomendasi).
       Format tabel contoh:
       | FLOW PROSES            | THREAT                    | C | I | Au | Av | N | SCENARIO                                                     | REKOMENDASI PENGAMANAN             |
       ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
       | contoh flow: user login | contoh threat: injection A06:01 | v |   | v  |   | v | fraudster injeksi form login                                 | implementasikan parameterized queries |

    4. Berikan penilaian DREAD (INFORMATIONAL=1, LOW=2, MEDIUM=3, HIGH=4, CRITICAL=5). Beri nilai masing-masing komponen dalam satu kolom, contoh: Discoverability=3, Reproducibility=2, dst. Rata-rata dari 5 komponen adalah skornya.

    5. Gunakan standar industri (OWASP, ASVS, NIST, Gartner) untuk analisis teknis.

    6. Batasi jawaban hanya dalam lingkup keamanan siber.

    7. Semua output Bahasa Indonesia dan jangan buang makna serapan Inggrisnya.
    """

    with st.sidebar:
        st.subheader(f"Selamat datang, {st.session_state['username']}!")

        if 'GROQ_API_KEY' not in st.session_state:
            st.session_state.GROQ_API_KEY = st.secrets.get("GROQ_API_KEY", "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").strip()

        st.subheader("Unggah File")
        st.markdown("Unggah file untuk dianalisis.")
        uploaded_file = st.file_uploader(
            "Pilih file",
            type=["jpeg", "jpg", "png", "pdf", "txt"]
        )

        st.subheader("Model")
        model_option = st.selectbox(
            'Pilih model yang akan digunakan:',
            ('llama-3.3-70b-versatile', 'qwen/qwen3-32b', 'deepseek-r1-distill-llama-70b')
        )

        if uploaded_file:
            st.success(f"File '{uploaded_file.name}' siap dianalisis.")

        if st.button("Logout"):
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            st.rerun()

    client = None
    if st.session_state.GROQ_API_KEY:
        try:
            client = groq.Groq(api_key=st.session_state.GROQ_API_KEY)
        except Exception as e:
            st.error(f"Gagal menginisialisasi klien Groq. Pastikan kunci API Anda valid. Error: {e}")
    else:
        st.warning("Kunci API Groq tidak ditemukan. Tambahkan di Secrets Streamlit: GROQ_API_KEY.")

    if "messages" not in st.session_state:
        st.session_state.messages = []

    # Tampilkan riwayat + unduhan
    for message in st.session_state.messages:
        with st.chat_message(message["role"]):
            st.markdown(message["content"])
            if message["role"] == "assistant":
                pdf_download_link = create_pdf_with_xhtml2pdf(
                    message["content"],
                    f"analisis_{message['content'][:20].replace(' ', '_')}.pdf"
                )
                st.markdown(pdf_download_link, unsafe_allow_html=True)
                excel_bytes = create_excel_from_markdown(
                    message["content"],
                    "alfa_threat_analysis.xlsx"
                )
                if excel_bytes:
                    st.download_button(
                        "📥 Unduh Analisis (Excel)",
                        data=excel_bytes,
                        file_name="alfa_threat_analysis.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

    if prompt := st.chat_input("Deskripsikan skenario atau ajukan pertanyaan keamanan..."):
        if not client:
            st.error("Tidak dapat melanjutkan. Klien Groq belum terinisialisasi.")
        else:
            with st.chat_message("user"):
                st.markdown(prompt)

            final_prompt = prompt
            if uploaded_file is not None:
                st.info(f"Menganalisis file: **{uploaded_file.name}**...")
                if uploaded_file.type in ["image/jpeg", "image/png"]:
                    image = Image.open(uploaded_file)
                    st.image(image, caption=f"File gambar: {uploaded_file.name}", use_column_width=True)
                final_prompt = f"Berdasarkan sebuah file bernama '{uploaded_file.name}', jawab pertanyaan ini dari sudut pandang keamanan siber: {prompt}"
                st.success("Konteks file berhasil ditambahkan ke dalam prompt.")

            st.session_state.messages.append({"role": "user", "content": final_prompt})

            try:
                system_message = {"role": "system", "content": SYSTEM_PROMPT_CONTENT}
                messages_to_send = [system_message] + st.session_state.messages

                with st.spinner("Alfa Threat sedang menganalisis... 🤔"):
                    chat_completion = client.chat.completions.create(
                        messages=messages_to_send,
                        model=model_option,
                    )
                    response_text = chat_completion.choices[0].message.content
                    st.session_state.messages.append({"role": "assistant", "content": response_text})

                    with st.chat_message("assistant"):
                        st.markdown(response_text)

                        pdf_download_link = create_pdf_with_xhtml2pdf(response_text)
                        st.markdown(pdf_download_link, unsafe_allow_html=True)

                        excel_bytes = create_excel_from_markdown(response_text, "alfa_threat_analysis.xlsx")
                        if excel_bytes:
                            st.download_button(
                                "📥 Unduh Analisis (Excel)",
                                data=excel_bytes,
                                file_name="alfa_threat_analysis.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

            except Exception as e:
                st.error(f"Terjadi kesalahan saat berkomunikasi dengan API Groq. Detail: {e}")
                if st.session_state.messages and st.session_state.messages[-1].get("role") == "user":
                    st.session_state.messages.pop()

# =========================
# 5) Login Page
# =========================
def login_page():
    st.title("Login Alfa Threat Model")
    st.write("Silakan masukkan kredensial Anda untuk melanjutkan.")
    VALID_CREDENTIALS = {
        "admin": "password123",
        "putri": "putri",
        "jeremy": "jeremy"
    }
    with st.form("login_form"):
        username = st.text_input("Username").lower()
        password = st.text_input("Password", type="password")
        submitted = st.form_submit_button("Login")
        if submitted:
            if username in VALID_CREDENTIALS and password == VALID_CREDENTIALS[username]:
                st.session_state.authenticated = True
                st.session_state.username = username
                st.session_state.last_activity = time.time()
                st.rerun()
            else:
                st.error("Username atau password salah.")

# =========================
# 6) App Flow & Session
# =========================
if 'authenticated' not in st.session_state:
    st.session_state.authenticated = False
if 'username' not in st.session_state:
    st.session_state.username = ""
if 'last_activity' not in st.session_state:
    st.session_state.last_activity = 0

# Auto-logout 30 menit idle
if st.session_state.authenticated:
    if time.time() - st.session_state.last_activity > 1800:
        for key in list(st.session_state.keys()):
            del st.session_state[key]
        st.warning("Sesi Anda telah berakhir karena tidak aktif. Silakan login kembali.")
        time.sleep(1.5)
        st.rerun()

if st.session_state.authenticated:
    main_app()
else:
    login_page()
